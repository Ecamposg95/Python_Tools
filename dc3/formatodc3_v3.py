import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import os
import logging

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Definir rutas
directorio_trabajo = r"C:\Users\ecamp\Devs\Python_Tools-main\dc3"
ruta_registros = os.path.join(directorio_trabajo, "DC3_resgistro.xlsx")
ruta_formato = os.path.join(directorio_trabajo, "DC-3-Formato2.xlsx")
carpeta_salida = os.path.join(directorio_trabajo, "Formatos_Generados")

if not os.path.exists(carpeta_salida):
    os.makedirs(carpeta_salida)

if not os.path.exists(ruta_registros):
    raise FileNotFoundError(f"No existe el archivo de registros: {ruta_registros}")
if not os.path.exists(ruta_formato):
    raise FileNotFoundError(f"No existe el archivo de formato: {ruta_formato}")

df = pd.read_excel(ruta_registros, sheet_name="Hoja1")

# Función para extraer cada dígito de una fecha
def descomponer_fecha(fecha):
    return [int(d) for d in fecha.strftime("%Y%m%d")]

# Función para fusionar dos celdas y escribir un solo valor centrado sin negritas
def escribir_doble_celda(ws, fila, col1, col2, valor):
    ws.merge_cells(start_row=fila, start_column=col1, end_row=fila, end_column=col2)
    celda = ws.cell(row=fila, column=col1, value=valor)
    celda.font = Font(bold=False)
    celda.alignment = Alignment(horizontal="center", vertical="center")

# Fuente común
font_normal = Font(bold=False)

# Iterar sobre los registros
for index, row in df.iterrows():
    try:
        wb = load_workbook(ruta_formato)
        ws = wb.active

        # Unmerge preventivo
        for merged in list(ws.merged_cells.ranges):
            if any(celda in merged.coord for celda in ["E17", "E29", "E38", "E41", "Z17"]):
                ws.unmerge_cells(str(merged.coord))

        # CURP (E17 en adelante)
        for i, char in enumerate(row["CURP"]):
            c = ws.cell(row=17, column=5 + i, value=char)
            c.font = font_normal
            c.alignment = Alignment(horizontal="center", vertical="center")

        # RFC (E29 en adelante)
        for i, char in enumerate(row["RFC_SHCP"]):
            c = ws.cell(row=29, column=5 + i, value=char)
            c.font = font_normal
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Ocupación (Z17)
        ws["Z17"] = row["Ocupacion"]
        ws["Z17"].font = font_normal
        ws["Z17"].alignment = Alignment(horizontal="left", vertical="center")

        # Nombre
        ws["E14"] = row["Nombre"]
        ws["E14"].font = font_normal
        ws["E14"].alignment = Alignment(horizontal="left", vertical="center")

        # Puesto, Empresa
        ws["E20"] = row["Puesto"]
        ws["E20"].font = font_normal

        ws["E26"] = row["Empresa"]
        ws["E26"].font = font_normal

        # Nombre del Curso
        ws["E35"] = row["N_Curso"]
        ws["E35"].font = font_normal

        # Duración del curso (E38:G38)
        ws.merge_cells("E38:G38")
        ws["E38"] = row["Duración del curso"]
        ws["E38"].font = font_normal
        ws["E38"].alignment = Alignment(horizontal="center", vertical="center")

        # Área Temática (E41:H41)
        ws.merge_cells("E41:H41")
        ws["E41"] = row["Tematica"]
        ws["E41"].font = font_normal
        ws["E41"].alignment = Alignment(horizontal="center", vertical="center")

        # Capacitador
        ws["E44"] = row["Capacitador"]
        ws["E44"].font = font_normal

        # Fechas
        fecha_ini = descomponer_fecha(row["F_inicial"])
        fecha_fin = descomponer_fecha(row["F_final"])

        # Año inicio (col 22 a 25)
        for i, col in enumerate([22, 23, 24, 25]):
            ws.cell(row=38, column=col, value=fecha_ini[i]).font = font_normal
            ws.cell(row=38, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Mes inicio (Z-AA y AB-AC)
        escribir_doble_celda(ws, 38, 26, 27, fecha_ini[4])
        escribir_doble_celda(ws, 38, 28, 29, fecha_ini[5])

        # Día inicio (AE-AF y AG-AH)
        escribir_doble_celda(ws, 38, 30, 31, fecha_ini[6])
        escribir_doble_celda(ws, 38, 32, 33, fecha_ini[7])

        # Año fin (col 36 a 39)
        for i, col in enumerate([36, 37, 38, 39]):
            ws.cell(row=38, column=col, value=fecha_fin[i]).font = font_normal
            ws.cell(row=38, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Mes fin (AJ-AK y AL-AM)
        escribir_doble_celda(ws, 38, 40, 41, fecha_fin[4])
        escribir_doble_celda(ws, 38, 42, 43, fecha_fin[5])

        # Día fin (AN-AO y AP-AQ)
        escribir_doble_celda(ws, 38, 44, 45, fecha_fin[6])
        escribir_doble_celda(ws, 38, 46, 47, fecha_fin[7])

        # Firmas
        ws["F53"] = row["Instructor"]
        ws["F53"].font = font_normal
        ws["U53"] = row["patron"]
        ws["U53"].font = font_normal
        ws["AH53"] = row["Representante"]
        ws["AH53"].font = font_normal

        # Guardar
        nombre_archivo = f"DC3_{row['Nombre'].replace(' ', '_')}.xlsx"
        ruta_guardado = os.path.join(carpeta_salida, nombre_archivo)
        wb.save(ruta_guardado)

        logging.info(f"✅ Formato generado: {ruta_guardado}")

    except Exception as e:
        logging.error(f"❌ Error en el registro {index + 1}: {e}")

logging.info("🎉 Todos los formatos han sido generados correctamente.")
