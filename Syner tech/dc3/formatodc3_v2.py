import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import os
import logging

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Definir rutas
directorio_trabajo = r"C:\Users\ecamp\Devs\Python_Tools\dc3"
ruta_registros = os.path.join(directorio_trabajo, "DC3_resgistro.xlsx")
ruta_formato = os.path.join(directorio_trabajo, "DC-3-Formato.xlsx")
carpeta_salida = os.path.join(directorio_trabajo, "Formatos_Generados")

# Crear la carpeta de salida si no existe
if not os.path.exists(carpeta_salida):
    os.makedirs(carpeta_salida)

# Verificar si los archivos de entrada existen
if not os.path.exists(ruta_registros):
    raise FileNotFoundError(f"El archivo de registros no existe: {ruta_registros}")
if not os.path.exists(ruta_formato):
    raise FileNotFoundError(f"El archivo de formato no existe: {ruta_formato}")

# Cargar base de datos de inscritos
try:
    df = pd.read_excel(ruta_registros, sheet_name="Hoja1")
except Exception as e:
    raise Exception(f"Error al leer el archivo de registros: {e}")

# Función para extraer cada dígito de una fecha en una lista
def descomponer_fecha(fecha):
    fecha_str = fecha.strftime("%Y%m%d")  # Convertir a string en formato YYYYMMDD
    return [int(digito) for digito in fecha_str]

# Iterar sobre cada inscrito
for index, row in df.iterrows():
    try:
        # Validar campos requeridos
        required_fields = ["Nombre", "CURP", "RFC_SHCP", "Ocupacion", "Duración del curso", "F_inicial", "F_final",
                           "Puesto", "Empresa", "N_Curso", "Tematica", "Capacitador", "Instructor", "patron", "Representante"]
        for field in required_fields:
            if pd.isna(row[field]):
                raise ValueError(f"El campo {field} está vacío en el registro {index + 1}")

        wb = load_workbook(ruta_formato)
        ws = wb.active

        # 🔹 Descombinar celdas si están combinadas
        for merged_range in ws.merged_cells.ranges:
            if "E17" in merged_range.coord:
                ws.unmerge_cells(merged_range.coord)
            if "E29" in merged_range.coord:
                ws.unmerge_cells(merged_range.coord)
            if "E38" in merged_range.coord:
                ws.unmerge_cells(merged_range.coord)
            if "E41" in merged_range.coord:
                ws.unmerge_cells(merged_range.coord)
            if "Z17" in merged_range.coord:
                ws.unmerge_cells(merged_range.coord)

        font_normal = Font(bold=False)

        # 🔹 CURP carácter por carácter
        for i, char in enumerate(row["CURP"]):
            cell = ws.cell(row=17, column=5 + i, value=char)
            cell.font = font_normal
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 🔹 RFC carácter por carácter
        for i, char in enumerate(row["RFC_SHCP"]):
            cell = ws.cell(row=29, column=5 + i, value=char)
            cell.font = font_normal
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 🔹 Ocupación en Z17 (una sola celda, alineado a la izquierda)
        ws["Z17"] = row["Ocupacion"]
        ws["Z17"].alignment = Alignment(horizontal="left", vertical="center")
        ws["Z17"].font = font_normal

        # 🔹 Nombre del trabajador (E14)
        ws["E14"] = row["Nombre"]
        ws["E14"].alignment = Alignment(horizontal="left", vertical="center")
        ws["E14"].font = font_normal

        # 🔹 Puesto (E20)
        ws["E20"] = row["Puesto"]
        ws["E20"].alignment = Alignment(horizontal="left", vertical="center")
        ws["E20"].font = font_normal

        # 🔹 Empresa (E26)
        ws["E26"] = row["Empresa"]
        ws["E26"].alignment = Alignment(horizontal="left", vertical="center")
        ws["E26"].font = font_normal

        # 🔹 Nombre del curso (E35)
        ws["E35"] = row["N_Curso"]
        ws["E35"].alignment = Alignment(horizontal="left", vertical="center")
        ws["E35"].font = font_normal

        # 🔹 Duración del curso (fusionar E38:G38)
        ws.merge_cells("E38:G38")
        ws["E38"] = row["Duración del curso"]
        ws["E38"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E38"].font = font_normal

        # 🔹 Área Temática (fusionar E41:H41)
        ws.merge_cells("E41:H41")
        ws["E41"] = row["Tematica"]
        ws["E41"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E41"].font = font_normal

        # 🔹 Capacitador (E44)
        ws["E44"] = row["Capacitador"]
        ws["E44"].alignment = Alignment(horizontal="left", vertical="center")
        ws["E44"].font = font_normal

        # 🔹 Fechas descompuestas
        digitos_inicio = descomponer_fecha(row["F_inicial"])
        digitos_fin = descomponer_fecha(row["F_final"])

        posiciones = {
            "año_inicio": [22, 23, 24, 25], "mes_inicio": [26, 27], "día_inicio": [30, 31],
            "año_fin": [36, 37, 38, 39], "mes_fin": [40, 41], "día_fin": [44, 45]
        }

        for i, col in enumerate(posiciones["año_inicio"]):
            ws.cell(row=38, column=col, value=digitos_inicio[i]).font = font_normal
        for i, col in enumerate(posiciones["mes_inicio"]):
            ws.cell(row=38, column=col, value=digitos_inicio[4 + i]).font = font_normal
        for i, col in enumerate(posiciones["día_inicio"]):
            ws.cell(row=38, column=col, value=digitos_inicio[6 + i]).font = font_normal
        for i, col in enumerate(posiciones["año_fin"]):
            ws.cell(row=38, column=col, value=digitos_fin[i]).font = font_normal
        for i, col in enumerate(posiciones["mes_fin"]):
            ws.cell(row=38, column=col, value=digitos_fin[4 + i]).font = font_normal
        for i, col in enumerate(posiciones["día_fin"]):
            ws.cell(row=38, column=col, value=digitos_fin[6 + i]).font = font_normal

        # 🔹 Firmas
        ws["H53"] = row["Instructor"]
        ws["H53"].alignment = Alignment(horizontal="left")
        ws["H53"].font = font_normal

        ws["U53"] = row["patron"]
        ws["U53"].alignment = Alignment(horizontal="left")
        ws["U53"].font = font_normal

        ws["AH53"] = row["Representante"]
        ws["AH53"].alignment = Alignment(horizontal="left")
        ws["AH53"].font = font_normal

        # 🔹 Guardar archivo generado
        nombre_archivo = f"DC3_{row['Nombre'].replace(' ', '_')}.xlsx"
        ruta_guardado = os.path.join(carpeta_salida, nombre_archivo)
        wb.save(ruta_guardado)

        logging.info(f"✅ Formato generado: {ruta_guardado}")
    except Exception as e:
        logging.error(f"❌ Error al procesar el registro {index + 1}: {e}")

logging.info("🎉 Proceso finalizado. Todos los formatos han sido generados correctamente.")


