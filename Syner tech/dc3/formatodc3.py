import pandas as pd
from openpyxl import load_workbook
import os

# Definir rutas
directorio_trabajo = r"C:\Users\ecamp\Devs\Python_Tools\dc3"
ruta_registros = os.path.join(directorio_trabajo, "DC3_resgistro.xlsx")
ruta_formato = os.path.join(directorio_trabajo, "DC-3-Formato.xlsx")
carpeta_salida = os.path.join(directorio_trabajo, "Formatos_Generados")

# Crear la carpeta de salida si no existe
if not os.path.exists(carpeta_salida):
    os.makedirs(carpeta_salida)

# Cargar base de datos de inscritos
df = pd.read_excel(ruta_registros, sheet_name="Hoja1")

# Función para extraer cada dígito de una fecha en una lista
def descomponer_fecha(fecha):
    fecha_str = fecha.strftime("%Y%m%d")  # Convertir a string en formato YYYYMMDD
    return [int(digito) for digito in fecha_str]  # Retornar los dígitos separados en una lista

# Iterar sobre cada inscrito
for index, row in df.iterrows():
    # Cargar la plantilla del Formato DC-3
    wb = load_workbook(ruta_formato)
    ws = wb.active

    # Extraer datos de la fecha
    digitos_inicio = descomponer_fecha(row["F_inicial"])
    digitos_fin = descomponer_fecha(row["F_final"])

    # 🔹 Verificar y descombinar celdas de la CURP, RFC y fechas si están combinadas
    for merged_range in ws.merged_cells.ranges:
        if "E17" in merged_range.coord:
            ws.unmerge_cells(merged_range.coord)
        if "E29" in merged_range.coord:
            ws.unmerge_cells(merged_range.coord)
        if "V38" in merged_range.coord:
            ws.unmerge_cells(merged_range.coord)

    # 🔹 Escribir la CURP carácter por carácter en la fila 17
    curp = row["CURP"]
    for i, char in enumerate(curp):
        ws.cell(row=17, column=5 + i, value=char)  # E17 en adelante

    # 🔹 Escribir el RFC carácter por carácter en la fila 29
    rfc = row["RFC_SHCP"]
    for i, char in enumerate(rfc):
        ws.cell(row=29, column=5 + i, value=char)  # E29 en adelante

    # 🔹 Escribir Año, Mes y Día (Inicio)
    ws.cell(row=38, column=22, value=digitos_inicio[0])  # V38 -> Año inicio (2)
    ws.cell(row=38, column=23, value=digitos_inicio[1])  # W38 -> Año inicio (0)
    ws.cell(row=38, column=24, value=digitos_inicio[2])  # X38 -> Año inicio (2)
    ws.cell(row=38, column=25, value=digitos_inicio[3])  # Y38 -> Año inicio (5)

    ws.cell(row=38, column=26, value=digitos_inicio[4])  # Z38 -> Mes inicio (0)
    ws.cell(row=38, column=27, value=digitos_inicio[5])  # AA38 -> Mes inicio (3)

    ws.cell(row=38, column=30, value=digitos_inicio[6])  # AD38 -> Día inicio (1)
    ws.cell(row=38, column=31, value=digitos_inicio[7])  # AE38 -> Día inicio (4)

    # 🔹 Escribir Año, Mes y Día (Fin)
    ws.cell(row=38, column=36, value=digitos_fin[0])  # AJ38 -> Año fin (2)
    ws.cell(row=38, column=37, value=digitos_fin[1])  # AK38 -> Año fin (0)
    ws.cell(row=38, column=38, value=digitos_fin[2])  # AL38 -> Año fin (2)
    ws.cell(row=38, column=39, value=digitos_fin[3])  # AM38 -> Año fin (5)

    ws.cell(row=38, column=40, value=digitos_fin[4])  # AN38 -> Mes fin (0)
    ws.cell(row=38, column=41, value=digitos_fin[5])  # AO38 -> Mes fin (3)

    ws.cell(row=38, column=44, value=digitos_fin[6])  # AR38 -> Día fin (1)
    ws.cell(row=38, column=45, value=digitos_fin[7])  # AS38 -> Día fin (4)

    # 🔹 Escribir otros datos en el formato
    ws.cell(row=14, column=5, value=row["Nombre"])  # E14 -> Nombre
    ws.cell(row=20, column=5, value=row["Puesto"])  # E20 -> Puesto
    ws.cell(row=26, column=5, value=row["Empresa"])  # E26 -> Empresa
    ws.cell(row=35, column=5, value=row["N_Curso"])  # E35 -> Nombre del Curso
    ws.cell(row=38, column=5, value=row["Duración del curso"])  # E38 -> Duración

    ws.cell(row=41, column=5, value=row["Tematica"])  # E41 -> Área Temática
    ws.cell(row=44, column=5, value=row["Capacitador"])  # E44 -> Agente Capacitador
    ws.cell(row=53, column=8, value=row["Instructor"])  # H53 -> Instructor
    ws.cell(row=53, column=21, value=row["patron"])  # U53 -> Patrón o Representante Legal
    ws.cell(row=53, column=34, value=row["Representante"])  # AH53 -> Representante de Trabajadores

    # Guardar con un nuevo nombre basado en el nombre del inscrito
    nombre_archivo = f"DC3_{row['Nombre'].replace(' ', '_')}.xlsx"
    ruta_guardado = os.path.join(carpeta_salida, nombre_archivo)
    wb.save(ruta_guardado)

    print(f"✅ Formato generado: {ruta_guardado}")

print("🎉 Proceso finalizado. Todos los formatos han sido generados correctamente.")
