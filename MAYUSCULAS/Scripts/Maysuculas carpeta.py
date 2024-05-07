# -*- coding: utf-8 -*-
"""
Created on Sun May  5 17:00:28 2024

@author: ecamp
"""

import openpyxl
import os

def convertir_a_mayusculas(ruta_entrada, ruta_salida):
    # Cargar el archivo de Excel
    workbook = openpyxl.load_workbook(ruta_entrada)
    
    # Iterar sobre todas las hojas del libro
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        
        # Iterar sobre todas las celdas de la hoja
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):  # Verificar si el valor de la celda es un string
                    cell.value = cell.value.upper()  # Convertir a mayúsculas

    # Guardar los cambios en la carpeta Output
    workbook.save(ruta_salida)

def procesar_archivos():
    # Cambiar el directorio de trabajo al directorio que contiene las carpetas Input y Output
    os.chdir('C:/Users/ecamp/SSCO/Herramientas Python/MAYUSCULAS')
    
    # Listar todos los archivos en la carpeta 'Input'
    archivos = os.listdir('Input')
    
    # Filtrar solo los archivos Excel (.xlsx)
    archivos_excel = [archivo for archivo in archivos if archivo.endswith('.xlsx')]
    
    # Procesar cada archivo
    for archivo in archivos_excel:
        ruta_entrada = os.path.join('Input', archivo)
        nombre_nuevo_archivo = archivo.replace('.xlsx', '_Mayus.xlsx')
        ruta_salida = os.path.join('Output', nombre_nuevo_archivo)
        convertir_a_mayusculas(ruta_entrada, ruta_salida)
        print(f'Archivo procesado: {nombre_nuevo_archivo}')

procesar_archivos()
