# -*- coding: utf-8 -*-
"""
Created on Sun May  5 17:08:30 2024

@author: ecamp
"""

import openpyxl
import os
from zipfile import BadZipFile

def convertir_a_mayusculas(ruta_entrada, ruta_salida):
    try:
        # Intentar cargar el archivo de Excel
        workbook = openpyxl.load_workbook(ruta_entrada)
        
        # Iterar sobre todas las hojas del libro
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            
            # Iterar sobre todas las celdas de la hoja
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):  # Verificar si el valor de la celda es un string
                        cell.value = cell.value.upper()  # Convertir a mayúsculas

        # Guardar los cambios en la carpeta Output
        workbook.save(ruta_salida)
        print(f'Archivo procesado exitosamente: {ruta_salida}')
    except BadZipFile:
        print(f'Error: El archivo {ruta_entrada} no pudo ser procesado como un archivo ZIP válido.')
    except Exception as e:
        print(f'Error desconocido con el archivo {ruta_entrada}: {str(e)}')

def procesar_archivos():
    os.chdir('C:/Users/ecamp/SSCO/Herramientas Python/MAYUSCULAS')
    
    archivos = os.listdir('Input')
    archivos_excel = [archivo for archivo in archivos if archivo.endswith('.xlsx')]
    
    for archivo in archivos_excel:
        ruta_entrada = os.path.join('Input', archivo)
        nombre_nuevo_archivo = archivo.replace('.xlsx', '_Mayus.xlsx')
        ruta_salida = os.path.join('Output', nombre_nuevo_archivo)
        convertir_a_mayusculas(ruta_entrada, ruta_salida)

procesar_archivos()
