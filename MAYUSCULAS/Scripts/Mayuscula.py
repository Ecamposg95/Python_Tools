# -*- coding: utf-8 -*-
"""
Created on Thu Apr  4 03:03:46 2024

@author: ecamp

SEPARA LOS NOMBRES ONCATENADOS DEL ARCHIVO ORIGINAL "2_Autor.xlsx"


"""
import openpyxl
import os

def convertir_a_mayusculas(nombre_archivo):
    # Cambiar el directorio de trabajo al directorio que contiene las carpetas Input y Output
    os.chdir('C:/Users/ecamp/SSCO/Herramientas Python/MAYUSCULAS')
    
    # Construir la ruta completa del archivo de entrada
    ruta_entrada = os.path.join('Input', nombre_archivo)
    # Cargar el archivo de Excel
    workbook = openpyxl.load_workbook(ruta_entrada)
    
    # Iterar sobre todas las hojas del libro
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        
        # Iterar sobre todas las celdas de la hoja
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):  # Verificar si el valor de la celda es un string
                    cell.value = cell.value.upper()  # Convertir a mayúsculas

    # Crear nuevo nombre para el archivo de salida añadiendo '_Mayus'
    nombre_nuevo_archivo = nombre_archivo.replace('.xlsx', '_Mayus.xlsx')
    ruta_salida = os.path.join('Output', nombre_nuevo_archivo)
    # Guardar los cambios en la carpeta Output
    workbook.save(ruta_salida)

# Nombre del archivo Excel en la carpeta Input
nombre_archivo = 'base de datos O.xlsx'
convertir_a_mayusculas(nombre_archivo)
